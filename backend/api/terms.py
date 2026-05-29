import csv
import io
import json
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile

from api.auth import get_current_admin
from database.mongo import get_database
from models.schemas import TermCreate, TermInDB, TermUpdate, UserInDB

router = APIRouter()


def object_id(term_id: str) -> ObjectId:
    try:
        return ObjectId(term_id)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Некорректный ID термина") from exc


@router.get("/", response_model=List[TermInDB])
async def get_terms(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=500),
    search: Optional[str] = None,
    category: Optional[str] = None,
    db=Depends(get_database),
):
    query = {}
    if search:
        query["$or"] = [
            {"ru": {"$regex": search, "$options": "i"}},
            {"kz": {"$regex": search, "$options": "i"}},
            {"en": {"$regex": search, "$options": "i"}},
            {"aliases": {"$regex": search, "$options": "i"}},
            {"deprecated": {"$regex": search, "$options": "i"}},
        ]
    if category:
        query["category"] = category

    cursor = db.terms.find(query).sort("ru", 1).skip(skip).limit(limit)
    return await cursor.to_list(length=limit)


@router.get("/stats")
async def get_term_stats(db=Depends(get_database)):
    total = await db.terms.count_documents({})
    categories = await db.terms.aggregate(
        [{"$group": {"_id": "$category", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}]
    ).to_list(length=100)
    with_english = await db.terms.count_documents({"en": {"$nin": [None, ""]}})
    return {"total": total, "with_english": with_english, "categories": categories}


@router.post("/", response_model=TermInDB)
async def create_term(term: TermCreate, db=Depends(get_database), admin: UserInDB = Depends(get_current_admin)):
    result = await db.terms.insert_one(term.model_dump())
    created_term = await db.terms.find_one({"_id": result.inserted_id})
    return TermInDB(**created_term)


@router.put("/{term_id}", response_model=TermInDB)
async def update_term(
    term_id: str,
    term: TermUpdate,
    db=Depends(get_database),
    admin: UserInDB = Depends(get_current_admin),
):
    payload = {key: value for key, value in term.model_dump().items() if value is not None}
    if not payload:
        raise HTTPException(status_code=400, detail="Нет данных для обновления")

    result = await db.terms.update_one({"_id": object_id(term_id)}, {"$set": payload})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Термин не найден")
    updated = await db.terms.find_one({"_id": object_id(term_id)})
    return TermInDB(**updated)


@router.delete("/{term_id}")
async def delete_term(term_id: str, db=Depends(get_database), admin: UserInDB = Depends(get_current_admin)):
    result = await db.terms.delete_one({"_id": object_id(term_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Термин не найден")
    return {"message": "Термин удален"}


@router.post("/import")
async def import_terms(
    file: UploadFile = File(...),
    replace: bool = Query(False),
    db=Depends(get_database),
    admin: UserInDB = Depends(get_current_admin),
):
    rows = await parse_terms_file(file)
    if not rows:
        raise HTTPException(status_code=400, detail="Файл не содержит терминов")

    if replace:
        await db.terms.delete_many({})

    inserted = 0
    updated = 0
    for row in rows:
        existing = await db.terms.find_one({"$or": [{"ru": row["ru"]}, {"kz": row["kz"]}]})
        if existing:
            await db.terms.update_one({"_id": existing["_id"]}, {"$set": row})
            updated += 1
        else:
            await db.terms.insert_one(row)
            inserted += 1

    return {"inserted": inserted, "updated": updated, "total": inserted + updated}


async def parse_terms_file(file: UploadFile) -> List[dict]:
    content = await file.read()
    name = (file.filename or "").lower()

    if name.endswith(".json"):
        data = json.loads(content.decode("utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("terms", [])
        return [normalize_row(row) for row in data if normalize_row(row)]

    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [normalize_row(row) for row in reader if normalize_row(row)]

    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="Для XLSX установите openpyxl") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(max_row=1))]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            raw = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
            normalized = normalize_row(raw)
            if normalized:
                rows.append(normalized)
        return rows

    raise HTTPException(status_code=400, detail="Поддерживаются CSV, JSON и XLSX")


def normalize_row(row: dict) -> Optional[dict]:
    kz = str(row.get("kz") or row.get("kk") or row.get("kazakh") or row.get("қазақша") or "").strip()
    ru = str(row.get("ru") or row.get("russian") or row.get("русский") or "").strip()
    en = str(row.get("en") or row.get("english") or "").strip() or None
    if not kz or not ru:
        return None

    aliases = row.get("aliases") or []
    deprecated = row.get("deprecated") or []
    if isinstance(aliases, str):
        aliases = [item.strip() for item in aliases.split(";") if item.strip()]
    if isinstance(deprecated, str):
        deprecated = [item.strip() for item in deprecated.split(";") if item.strip()]

    return {
        "kz": kz,
        "ru": ru,
        "en": en,
        "category": str(row.get("category") or row.get("категория") or "Фармацевтика").strip(),
        "aliases": aliases,
        "deprecated": deprecated,
        "notes": str(row.get("notes") or row.get("примечание") or "").strip() or None,
    }
